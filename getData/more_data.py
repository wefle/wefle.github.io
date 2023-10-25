import xlrd
import numpy as np
import requests 
from geopy.geocoders import Nominatim, ArcGIS
from bs4 import BeautifulSoup 
import re
from openpyxl import load_workbook

ORI_FILE = "output/stempel.xlsx"

#open excel file
work_book = load_workbook(ORI_FILE)
sheets = work_book.sheetnames
stamps = work_book[sheets[0]]
print("opened file")

def check_data_for_missing_pos(raw_data):
    data = []

    #filter for names where no positions where written 
    for i in range(0, len(raw_data[1])):
        if(raw_data[1][i].value == ""):
            data.append([i, raw_data[0][i].value])

    return data

def get_raw_data(file):
    book = xlrd.open_workbook(file, encoding_override = "utf-8")
    sheet = book.sheet_by_index(0)

    #get names and positions
    name_data = np.asarray([sheet.cell(i, 0) for i in range(1, sheet.nrows)])
    pos_data = np.asarray([sheet.cell(i, 1) for i in range(1, sheet.nrows)])

    return name_data, pos_data

def find_and_add_postions(data):

    lat = ""
    lng = ""

    for elem in data:
        address = elem[1]
        #went somewhere nowhere
        if(elem[1].find("KUGELsPASS") != -1):
                address = "Schlosslinde Harzgerode"

        geolocator = Nominatim(user_agent="granit_stein")
        location = geolocator.geocode(address, language="de", exactly_one=True)  

        if(location == None):
            #extra cases because nomination couldnt find them :(
            if(elem[1].find(" in ") != -1):
                address = elem[1].replace(" in ", ",") 
            elif(elem[1].find("Hüttengarten") != -1):
                address = elem[1].replace("Hüttengarten", "Cafe") 
            elif(elem[1].find("Seilhängebrücke") != -1): 
                address = elem[1].replace("Seilhängebrücke", "Harz")
            elif(elem[1].find("Transall-Sonderstempel ") != -1): 
                address = elem[1].replace("Transall-Sonderstempel ", "")
            elif(elem[1].find("Waldlehrpfad") != -1):    
                address = elem[1].replace("Waldlehrpfad", "Waldschule")  
            elif(elem[1].find("Märchenweg") != -1):    
                address = elem[1].replace("Märchenweg", "Kapitelsberg 15B")  
            elif(elem[1].find("HSB") != -1):     
                address = elem[1].replace("HSB", "Aussichtspunkt") 
            elif(elem[1].find("Bergkönigin") != -1):  
                address = elem[1].replace("Bergkönigin", "") 
            elif(elem[1].find("Auerhuhn") != -1):  
                address = elem[1].replace("Auerhuhngehege", "Auerhuhn-Schaugehege")     
            elif(elem[1].find("Bienenpfad") != -1):  
                address = elem[1].replace("Bienenpfad", "Glockenturm")
            elif(elem[1].find("Bundeswehr") != -1):  
                address = elem[1].replace("Bundeswehr Ausstellung, ", "")       
            location = geolocator.geocode(address, language="de", exactly_one=True)

        if(location != None):
            lat = str(round(location.latitude, 5)) + " | "
            lng = str(round(location.longitude, 5)) + " | "    
        
        if(lat != "" and lng != ""):
            #+2 -> first row + starting at 1 in file (data array started at 0)
            save_pos([lat, lng], elem[0]+2)
            lat = lng = ""

def save_pos(pos, r):   
    #longitude and latitude
    stamps.cell(row = r, column = 2).value = pos[0]
    stamps.cell(row = r, column = 3).value = pos[1]        

def remove_dups(arr):
    #check if 2dim or not
    #2dim -> sightseeings + links
    if len(arr[0]) == 2:
        start = arr[0][0]
        i = 1
        while i < len(arr):
            if (arr[i][0].find(start) != -1  and len(arr[i][0]) == len(start)): 
                arr.pop(i)
                i -= 1
            if arr[i][0].startswith("Baumwipfelpfad"):
                if start.startswith("Baumwipfelpfad"):
                    arr.pop(i)
                    i -= 1
                else:
                    start = arr[i][0]
            if arr[i][0].startswith("Kyffhäuserdenkmal"):
                arr.pop(i)
                i -= 1
            else: 
                start = arr[i][0]
            i += 1
    #1dim -> websites
    else:   
        start = arr[0]
        i = 1
        while i < len(arr):
            if arr[i].find(start) != -1 :
                arr.pop(i)
                i -= 1
            else: 
                start = arr[i]
            i += 1

    return (arr)

def dim(a):
    if not type(a) == list:
        return []
    return [len(a)] + dim(a[0])

sight = None
def filter_function(elem):
    global sight

    #hexentanzplatz

    if sight != None:

        #prepare string lists to find matching words
        #try removing or changing unnessessary words or 
        #words with semantic similiar value 
        #(klostergrund = kloster, burgruine = burg + ruine)
        
        stopword = ['in', 'im', 'von', '','die', 'harz']

        e_s = elem.value.lower()
            #delete special chars
        e_s = re.sub(r'[()]|\W+', ' ', e_s)

        if e_s.find("bad") != -1: e_s = re.sub(r"(bad )", "bad", e_s)
        if e_s.find("burgruine") != -1: e_s = e_s.replace("burgruine", "burg ruine")
        if e_s.find("kloster") != -1: e_s = e_s.replace("klostergrund", "kloster")

        e_s = e_s.split(" ")
        e_s = [i for i in e_s if i not in stopword]

        sight_s = sight[0].lower()
        sight_s = re.sub(r'[()]|\W+', ' ', sight_s)

        if sight_s.find("bad") != -1: sight_s = re.sub(r"(bad )", "bad", sight_s)
        if sight_s.find("burgruine") != -1: sight_s = sight_s.replace("burgruine", "burg ruine")
        if sight_s.find("kloster") != -1: sight_s = sight_s.replace("klostergrund", "kloster")
        if sight_s.find("schnarcher") != -1: sight_s = sight_s.replace("schnarcherklippen", "schnarcherklippe")

        sight_s = sight_s.split(" ")
        sight_s = [i for i in sight_s if i not in stopword]
    
        matches_num = len(list(set(e_s)&set(sight_s)))

        if matches_num > 0:

            #exact one word and matching
            if len(e_s) == matches_num:
                return True
            
            #more words
            if len(e_s) > 1:
                ratio = len(e_s) - matches_num

                #same amount of matching and unmatching words 
                if ratio == len(e_s) / 2: 
                    
                    if len(e_s) == len(sight_s):
                        e_s = " ".join(e_s)
                        sight_s = " ".join(sight_s)
                        if e_s == sight_s:
                            return True
                        else:
                            return False
                         
                    if len(e_s) > len(sight_s):
                        length = round(len(e_s) / 2)
                        e = [" ".join(e_s[:length])] + [" ".join(e_s[length:])]
                        e = [i for i in e if i != ""]
                        s = [" ".join(sight_s[:len(e_s) - length])] + [" ".join(sight_s[len(e_s)-length:])]
                        s = [i for i in s if i != ""]
                        if e[-1] == s[-1]:
                            return True
                        elif e[0] == s[0]:
                            return True
                        else:
                            return False
                        
                    else:
                        length = round(len(sight_s) / 2)
                        s = [" ".join(sight_s[:length])] + [" ".join(sight_s[length:])]
                        s = [i for i in s if i != ""]
                        e = [" ".join(e_s[:len(sight_s)-length])] + [" ".join(e_s[len(sight_s)-length:])]
                        e = [i for i in e if i != ""]
                        if e[-1] == s[-1]:
                            return True
                        elif e[0] == s[0]:
                            return True
                        else:
                            return False

                #more unmatching words than matching ones 
                if ratio > len(e_s) / 2:
                    #Glasmanufaktur = special
                    if(sight[0].find("Glasmanufaktur") != -1 and
                       elem.value.find("Glasmanufaktur") != -1):
                        return True
                    return False 
                
                #more matching words than unmatching ones   
                else:
                    #special cases: burg, ruine, burgruine
                    if e_s[0] == sight_s[0]:
                        if e_s[-1] == sight_s[-1]:
                            return True  
                        else: 
                            return False
                    return True             
        
        #no matching words
        else:
            #National = special
            if(sight[0].find("National") != -1 and
               elem.value.find("National") != -1):
                return True
            #Brocken = special
            if(sight[0].startswith("Brocken") and
               elem.value.startswith("Brocken")):
                return True
            return False

def save_sightseeing(data, elem, trivia):

    if len(elem) > 1 :
        for e in elem:
            r = np.where(data == e)[0][0] + 2
            #sightseeing
            stamps.cell(row = r, column = 5).value = "ja"
            #trivia
            if stamps.cell(row = r, column = 6).value != None:
                stamps.cell(row = r, column = 6).value += " | " + str(trivia)
            else:
                stamps.cell(row = r, column = 6).value = str(trivia)
    else: 
        r = np.where(data == elem)[0][0] + 2
        #sightseeing
        stamps.cell(row = r, column = 5).value = "ja"
        #trivia
        if stamps.cell(row = r, column = 6).value != None:
                stamps.cell(row = r, column = 6).value += " | " + str(trivia)
        else:
            stamps.cell(row = r, column = 6).value = str(trivia)

def find_and_add_sightseeings(data):
    global sight
    #search for same names in data -> nein => ja
    #add to file
    try:
        from googlesearch import search
    except ImportError: 
        print("No module named 'google' found")
    
    #what to search for
    query = "Harz Sehenswürdigkeiten"
    queries = []
    
    #get websites of interesst
    for website in search(query, tld="co.in", num=10, stop=10, pause=2):
        queries.append(website)
    
    #remove duplicates
    if queries:
        queries = remove_dups(queries)
        
    sightseeings = []
    
    for elem in queries:
        #get website for info searching
        resp = requests.get(elem)

        #http_respone 200 means OK status 
        if resp.status_code == 200:
            print("Successfully opened the web page") 

            soup = BeautifulSoup(resp.text, "html.parser")
            if(elem.find("Regionales") != -1):
                content = soup.find("div", {"class":"regional-box-main"})            
                for e in content.findAll('h3'):
                    if e.text.find("ß") != -1: s = e.text.replace("ß", "ss")
                    else: s = e.text
                    sightseeings.append([re.sub(r'\W', ' ', s), [e.find_next("a"), elem]])
            elif(elem.find("voucher") != -1):
                content = soup.find("div", {"class":"entry-content"})           
                for e in content.findAll('h2'):
                    s = re.sub(r'\W', ' ', e.text)
                    s = re.sub(r'[0-9]', '', s)
                    if s.startswith(' '):
                        s = s.replace('  ', '', 1)
                    sightseeings.append([s, [e.find_next("a"), elem]])
            else:
                content = soup.find("div", {"class":"blogArticleContent"})            
                for e in content.findAll('h3'):
                    s = re.sub(r'\W', ' ', e.text)
                    s = re.sub(r'[0-9]', '', s)
                    if s.startswith(' '):
                        s = s.replace('  ', '', 1)
                    sightseeings.append([s, [e['id'], elem]])
        else:
            print("Error: I couldn't open the website. :'(")
    
    sightseeings = sorted(sightseeings, key= lambda elem: elem[0].lower())
    sightseeings = remove_dups(sightseeings) 
    
    while sightseeings:
        #found -> save, delete
        #not found -> delete
        sight = sightseeings[0]        
        elem = list(filter(filter_function, data))
        if elem: 
            save_sightseeing(data, elem, sight[1])        
        sightseeings.pop(0) 
       

raw_data = get_raw_data(ORI_FILE)
data = check_data_for_missing_pos(raw_data)
find_and_add_postions(data)
find_and_add_sightseeings(raw_data[0])

#close and save excel file
work_book.save(ORI_FILE)
print("saved file")

